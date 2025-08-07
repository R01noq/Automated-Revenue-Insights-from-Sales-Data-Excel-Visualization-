# Importing necessary libraries
import pandas as pd 
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Load the dataset from CSV file
dt = pd.read_csv("D:\\my python\\performance_report.csv")
print(dt)

# Calculate missing values percentage
missig_rateo = dt.isnull().mean()*100
print(missig_rateo)

# Map full month names to short names
month_map = {
    "January": "Jan", "February": "Feb", "March": "Mar",
    "April": "Apr", "May": "May", "June": "Jun",
    "July": "Jul", "August": "Aug", "September": "Sep",
    "October": "Oct", "November": "Nov", "December": "Dec"
}
dt["Month"] = dt["Month"].map(month_map)

# Define the correct month order
dt["Month"] = pd.Categorical(dt["Month"], categories=["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], ordered=True)

# Convert revenue, cost, profit, and margin columns to numeric
cols = ["Revenue", "Cost", "Profit", "Margin(%)"]
for col in cols:
    dt[col] = pd.to_numeric(dt[col], errors="coerce")

# Group by Month and Product
rev_product = dt.groupby(["Month", "Product"]).agg(
    Total_Revenue=("Revenue", "sum"),
    Total_Profit=("Profit", "sum")
).reset_index()

# Find best and worst product by revenue
best_product = rev_product.loc[rev_product.groupby("Product")["Total_Revenue"].idxmax()].sort_values(by="Total_Revenue", ascending=False).head(1)
worst_product = rev_product.loc[rev_product.groupby("Product")["Total_Revenue"].idxmin()].sort_values(by="Total_Revenue", ascending=True).head(1)

# Group by Month and Branch
rev_branch = dt.groupby(["Month", "Branch"]).agg(
    Total_Revenue=("Revenue", "sum"),
    Total_Profit=("Profit", "sum")
).reset_index()

# Find best and worst branch by revenue
best_branch = rev_branch.loc[rev_branch.groupby("Branch")["Total_Revenue"].idxmax()].sort_values(by="Total_Revenue", ascending=False).head(1)
worst_branch = rev_branch.loc[rev_branch.groupby("Branch")["Total_Revenue"].idxmin()].sort_values(by="Total_Revenue", ascending=True).head(1)

# Group by Month only for total revenue and profit
df = dt.groupby("Month").agg(
    Total_Revenue=("Revenue", "sum"),
    Total_Profit=("Profit", "sum")
).reset_index()

# Find best and worst month
best_month = df.loc[df.groupby("Month")["Total_Revenue"].idxmax()].sort_values(by="Total_Revenue", ascending=False).head(1)
worst_month = df.loc[df.groupby("Month")["Total_Revenue"].idxmin()].sort_values(by="Total_Revenue", ascending=True).head(1)

# Visualize product performance over months
plt.figure(figsize=(12,6))
plt.subplot(2,1,1)
sns.lineplot(data=rev_product, x="Month", y="Total_Revenue", hue="Product", marker="o")
plt.title("Revenue per Product over Months")
plt.xlabel("Month")
plt.ylabel("Revenue")
plt.legend(loc='lower left', bbox_to_anchor=(1, 0.5))
plt.grid(True)
plt.tight_layout()

plt.subplot(2,1,2)
sns.lineplot(data=rev_product, x="Month", y="Total_Profit", hue="Product", marker="o")
plt.title("profits per Product over Months")
plt.xlabel("Month")
plt.ylabel("Revenue")
plt.legend(loc='lower left', bbox_to_anchor=(1, 0.5))
plt.grid(True)
plt.tight_layout()
plt.savefig("monthly_product_analyist.png", bbox_inches="tight", dpi=300)
plt.show()

# Visualize branch performance over months
plt.figure(figsize=(12,6))
plt.subplot(2,1,1)
sns.barplot(data=rev_branch, x="Month", y="Total_Revenue", hue="Branch", edgecolor="black")
plt.title("Monthly revenue for each branch ")
plt.xlabel("Month")
plt.ylabel("Revenue")
plt.xticks(rotation=45)
plt.legend(loc="lower left", bbox_to_anchor=(1, 0.5))
plt.tight_layout()

plt.subplot(2,1,2)
sns.barplot(data=rev_branch, x="Month", y="Total_Profit", hue="Branch", edgecolor="black")
plt.title("Monthly profit for each branch ")
plt.xlabel("Month")
plt.ylabel("Profit")
plt.xticks(rotation=45)
plt.legend(loc="lower left", bbox_to_anchor=(1, 0.5))
plt.tight_layout()
plt.savefig("monthly_branch_analyist.png", bbox_inches="tight", dpi=300)
plt.show()

# Compare overall monthly revenue vs profit
n = len(df)
x = np.arange(n)
width=0.3
plt.figure(figsize=(12,6))
plt.bar(x+width/2, df["Total_Revenue"], label="Revenue", width=width, color="orange", edgecolor="black")
plt.bar(x-width/2, df["Total_Profit"], label="Profit", width=width, color="skyblue", edgecolor="black")
plt.title("comparison between (revenues/profits) per month ")
plt.xlabel("Month")
plt.ylabel("Revenue/Profit")
plt.xticks(ticks=range(n), labels=df["Month"])
plt.legend(loc="upper right")
plt.grid(True)
plt.tight_layout()
plt.savefig("monthly_analyist.png", bbox_inches="tight", dpi=300)
plt.show()

# Create Excel workbook
wb = Workbook()

# Sheet 1 - Product Table
ws1 = wb.create_sheet("Product")
for r in dataframe_to_rows(rev_product[["Month", "Product", "Total_Revenue", "Total_Profit"]],index=False, header=True):
    ws1.append(r)

# Sheet 2 - Product Plot
ws1_plot = wb.create_sheet("Plots 1")
img_1 = Image("monthly_product_analyist.png") 
ws1_plot.add_image(img_1, "A5") 

# Sheet 3 - Product Report
summary_ws1 = wb.create_sheet("Product Report")
report_product = [
    "Through my analysis of the data I find :",
    "",
    "First: Analyze the revenue of each product per month :",
    "",
    "1. The best product is (Tablet) with revenue value (175894) and net profit (48892.99) and that was in the month of (October).",
    "",
    "2. The worst product is (Printer) with revenue value (74225) and net profit (21513.98) and that was in the month of (May).",
    "",
]
for i,line in enumerate(report_product, start=1):
    summary_ws1[f"A{i}"] = line
summary_ws1.column_dimensions[get_column_letter(1)].width = 100

# Sheet 4 - Branch Table
ws2 = wb.create_sheet("Branch")
for r in dataframe_to_rows(rev_branch[["Month", "Branch", "Total_Revenue", "Total_Profit"]],index=False, header=True):
    ws2.append(r)

# Sheet 5 - Branch Plot
ws2_plot = wb.create_sheet("plots 2")
img_2 = Image("monthly_branch_analyist.png") 
ws2_plot.add_image(img_2, "A5")  

# Sheet 6 - Branch Report
summary_ws2 = wb.create_sheet("Branch report")
report_branch = [
    "Second: Analyze the revenue of each brunch per month :",
    "",
    "1. The best branch is (East) with revenue value (204983) and net profit (53157.93) and that was in the month of (November).",
    "",
    "2. The worst branch is (North) with revenue value (77109) and net profit (24821.59) and that was in the month of (December)."
]
for i,line in enumerate(report_branch, start=1):
    summary_ws2[f"A{i}"] = line
summary_ws2.column_dimensions[get_column_letter(1)].width = 100

# Sheet 7 - Monthly Table
ws3 = wb.create_sheet("Monthly")
for r in dataframe_to_rows(df[["Month", "Total_Revenue", "Total_Profit"]], index=False, header=True):
    ws3.append(r)

# Sheet 8 - Monthly Plot
ws3_plot = wb.create_sheet("Plots 3")
img_3 = Image("monthly_analyist.png")
ws3_plot.add_image(img_3, "A5")

# Sheet 9 - Monthly Report
summary_ws3 = wb.create_sheet("Monthly Report")
monthly_report = [
    "Third: Analyze the monthly revenue :",
    "",
    "1. The best branch is (October) with revenue value (690235) and net profit (183192.4).",
    "",
    "2. The worst branch is (May) with revenue value (493184) and net profit (130197.62).",
    "",
    "3. We notice a stable revenue growth during the period from January to April with revenue value (about: 55,000).",
    "",
    "4. From May to June, we noticed a (10%) in revenues and profit.",
    "",
    "5. During the month of July, revenues and profits increased by (20%) compared to the previous two months.",
    "",
    "6. During the months of August and September, we noticed stability in the value of revenues, with a (10%) decrease compared to July.",
    "",
    "7. During the month of October, revenues skyrocket by nearly (30%) about (70,000).",
    "",
    "8. Revenues drop by (6%) in November and continue to decline until they reach 52,000 in December."
]
for i, line in enumerate(monthly_report, start=1):
    summary_ws3[f"A{i}"] = line
summary_ws3.column_dimensions[get_column_letter(1)].width = 100

# Save the workbook
wb.save("product.xlsx")